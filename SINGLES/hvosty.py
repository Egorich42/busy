import sqlite3

import os
import openpyxl
from openpyxl import load_workbook,Workbook
from forge import create_list_of_table_values, grouping_by_key
from forge import sel_out_pays_br, sel_out_docs_br, sel_out_docs_br_back, sel_income_docs_br, sel_income_pays_br, sel_income_pays_br_dpd, sel_income_pays_br_back 
BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'



class Hvosty:
	def __init__(self, base_name=None, start_data = None, end_data = None):
		self.base_name = base_name
		self.start_data = start_data
		self.end_data = end_data

	def get_ops_list(self):
		conn = sqlite3.connect(self.base_name )
		cur = conn.cursor()

		income_pays_list = create_list_of_table_values(cur.execute(sel_income_pays_br.format(self.start_data, self.end_data)),cur.description)
		out_docs_list = create_list_of_table_values(cur.execute(sel_out_docs_br.format(self.start_data, self.end_data)),cur.description)

		out_pays_list = create_list_of_table_values(cur.execute(sel_out_pays_br.format(self.start_data, self.end_data)),cur.description)
		income_docs_list = create_list_of_table_values(cur.execute(sel_income_docs_br.format(self.start_data, self.end_data)),cur.description)

		conn.commit()
		conn.close()
		return income_pays_list, out_docs_list, out_pays_list, income_docs_list
		pass

	def contragent_ops_result(self, income_list):
		result = []
		for contragent in income_list:
			result +=[{'name': contragent[0]['contragent_name'], 'parent':contragent[0]['parent'],  'sum':round(sum([float(x['summ']) for x in contragent]),2)}]
		return result
		pass
		

	def found_result(self, one_list, sec_list):
		out_list = []
		in_list = []
		for doc in one_list:
			for ops in sec_list:
				if doc['parent'] == ops['parent']:
					if doc['sum'] > ops['sum']:
						out_list+=[{'name':doc['name'], 'summ':round(doc['sum'] -  ops['sum'],2)}]

					if doc['sum'] < ops['sum']:
						in_list+=[{'name':doc['name'], 'summ': round(ops['sum'] - doc['sum'],2)}]

		return(out_list, in_list)			
		pass


	def show_contragent_balance(self):
#		provider_debt = []#предоплатили поставщику, не отгрузил
#		provider_prepay = []#поставил, не оплатили еще

#		buyer_debt = []#покупатель должен денег
#		buyer_prepay = []#покупатель предоплатил, не отгрузили

		out_pays = self.contragent_ops_result(grouping_by_key(self.get_ops_list()[2],'parent'))
		income_docs= self.contragent_ops_result(grouping_by_key(self.get_ops_list()[3], 'parent'))

		income_pays = self.contragent_ops_result(grouping_by_key(self.get_ops_list()[0], 'parent'))
		outcome_docs= self.contragent_ops_result(grouping_by_key(self.get_ops_list()[1], 'parent'))

		provider_debt = self.found_result(out_pays, income_docs)[0]
		provider_prepay = self.found_result(out_pays, income_docs)[1]

		buyer_debt =self.found_result(income_pays, outcome_docs)[0]
		buyer_prepay = self.found_result(income_pays, outcome_docs)[1]

		return provider_debt, provider_prepay, buyer_debt, buyer_prepay
		pass


	def create_hvosty_excel(self):
		output_doc = BASE_DIR+'\\'+"resultat.xlsx"
		openpyxl.Workbook().save(output_doc)
		output_list = load_workbook(output_doc,data_only = True)
		main_out_sheet = output_list.active


		def insert_cell(row_val, col_val, cell_value):
			main_out_sheet.cell(row = row_val, column = col_val).value = cell_value
			pass

		insert_cell(1, 1, "Контрагент")
		insert_cell(1, 2, "Сумма")

		first_space = len(self.show_contragent_balance()[0])+12
		sec_space =  first_space++len(self.show_contragent_balance()[1])+6
		thr_space = sec_space+len(self.show_contragent_balance()[2])+6

		insert_cell(3, 1, 'ЗАДОЛЖЕННОСТЬ ПОКУПАТЕЛЕЙ (Д 62)'  )
		for i in range(len(self.show_contragent_balance()[0])):
			insert_cell(i+5, 1, self.show_contragent_balance()[0][i]['name'] )
			insert_cell(i+5, 2, self.show_contragent_balance()[0][i]['summ'])


		insert_cell(first_space, 1, 'АВАНСЫ ПОКУПАТЕЛЕЙ (КР 62)'  )
		for i in range(len(self.show_contragent_balance()[1])):
			insert_cell(i+first_space+2, 1, self.show_contragent_balance()[1][i]['name'] )
			insert_cell(i+first_space+2, 2, self.show_contragent_balance()[1][i]['summ'] )    

		insert_cell(sec_space, 1, 'АВАНСЫ ПОСТАВЩИКАМ(Д 60)' )
		for i in range(len(self.show_contragent_balance()[2])):
			insert_cell(i+sec_space+2, 1, self.show_contragent_balance()[2][i]['name'] )
			insert_cell(i+sec_space+2, 2, self.show_contragent_balance()[2][i]['summ'] )  
		
		insert_cell(thr_space, 1, 'ЗАДОЛЖЕННОСТЬ ПЕРЕД ПОСТАВЩИКАМИ(КР 60)' )
		for i in range(len(self.show_contragent_balance()[3])):
			insert_cell(i+thr_space+2, 1, self.show_contragent_balance()[3][i]['name'] )
			insert_cell(i+thr_space+2, 2, self.show_contragent_balance()[3][i]['summ'] ) 
		
		output_list.save(filename = output_doc)

		return(str(output_doc))
		pass






balance_lists = Hvosty("avangard.sqlite","'"+ '2016-06-30'+"'", "'"+'2018-03-30'+"'").create_hvosty_excel()



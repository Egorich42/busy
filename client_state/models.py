#! /usr/bin/env python
# -*- coding: utf-8 -*
from django.db import migrations, models
from django.http import HttpResponse, HttpResponseRedirect
from django.core.urlresolvers import reverse

import datetime
from datetime import date
import os
import sqlite3


from collections import defaultdict
from operator import itemgetter
import itertools
from itertools import groupby
import requests
import openpyxl
from openpyxl import load_workbook,Workbook


from forge import nbrb_rates_today, nbrb_rates_on_date, rates, years, months, days
from forge import create_list_of_table_values, sum_of_list, return_excel_list, generate_data_list, grouping_by_key
from forge.requests import *


BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'
TO_BASE_PATH = str('\\'.join(BASE_DIR.split('\\')[:-2]))+'\\'
TO_DOCS_PATH = str('\\'.join(BASE_DIR.split('\\')[:-2]))+'\\'+'media'+'\\'+'docs'+'\\'


def create_sorted_list(income_list):
    output_list = []
    full_grouped_list = []

    sorted_list = sorted(income_list, key=itemgetter('name'))

    for key, group in itertools.groupby(sorted_list, key=lambda x:x['name']):
        grouped_sorted_list = list(group)
        full_grouped_list += [grouped_sorted_list]                              

    for i in full_grouped_list:
        output_list+=[{'name': i[0]['name'],
                       'unp':str(i[0]['unp']), 
                       'nds':round(sum([float(x['nds']) for x in i]),2)}]

    return output_list
    pass
 
#---------------------------------CURENT FIN STATE---------------------------------#

class CompanyBalance:
    def __init__(self, base_name = None, data_start = None, data_end = None):
        self.base_name = base_name
        self.data_start = data_start
        self.data_end = data_end


    def tax_sum(self, select_command, cursor, tax_type):
        return sum([float(x[tax_type]) for x in create_list_of_table_values(cursor.execute(select_command.format(self.data_start, self.data_end)),cursor.description) if x[tax_type] != None])


    def count_nds(self):
        conn = sqlite3.connect(self.base_name)
        cur = conn.cursor()
        outcome_nds_sum = self.tax_sum(outcome_serv_nds, cur, 'nds')+self.tax_sum(outcome_tn_nds, cur, 'nds')
        income_nds_sum = self.tax_sum(income_serv_nds, cur, 'nds')+self.tax_sum(income_tn_nds, cur, 'nds')+self.tax_sum(income_tovary, cur, 'nds')
 
        conn.commit()
        conn.close()        

        return (outcome_nds_sum, income_nds_sum, outcome_nds_sum - income_nds_sum)
        pass


    def count_usn(self):
        conn = sqlite3.connect(self.base_name)
        cur = conn.cursor()
        usn_sum = round(self.tax_sum(outcome_full_nonds, cur, 'summ')*0.05, 2)
 
        conn.commit()
        conn.close()

        return usn_sum
        pass


    def create_tax_excel(request, income_list, tax_type):
        output_doc = BASE_DIR+'\\'+"tax_result.xlsx"
        openpyxl.Workbook().save(output_doc)
        output_list = load_workbook(output_doc,data_only = True)
        main_out_sheet = output_list.active    

        def insert_cell(row_val, col_val, cell_value):
            main_out_sheet.cell(row = row_val, column = col_val).value = cell_value
            pass


        if tax_type == "nds":
            insert_cell(1, 1, "НДС за выбранный период составляет")
            insert_cell(2, 1, "Входящий НДС")
            insert_cell(3, 1, "Исходящий НДС")
            insert_cell(4, 1, "НДС к уплате:")
            insert_cell(2, 2, str(round(income_list[0],2))+" "+"руб.")
            insert_cell(3, 2, str(round(income_list[1],2))+" "+"руб.")
            insert_cell(4, 2, str(round(income_list[2],2))+" "+"руб.")

        
        else: 
            insert_cell(1, 1, "УСН за выбранный период составляет")  
            insert_cell(1, 2, str(round(income_list,2))+" "+"руб.") 


        output_list.save(filename = output_doc)

        return(str(output_doc))
        pass


#---------------------------------HVOSTY---------------------------------#

class Hvosty:
    def __init__(self, base_name=None, start_data = None, end_data = None):
        self.base_name = base_name
        self.start_data = start_data
        self.end_data = end_data

    def get_ops_list(self):
        conn = sqlite3.connect(self.base_name )
        cur = conn.cursor()

        income_pays_list = create_list_of_table_values(cur.execute(sel_income_pays_br.format(self.start_data, self.end_data)),cur.description)
        out_docs_list = create_list_of_table_values(cur.execute(sel_out_docs_br.format(self.start_data, self.end_data)),cur.description)

        out_pays_list = create_list_of_table_values(cur.execute(sel_out_pays_br.format(self.start_data, self.end_data)),cur.description)
        income_docs_list = create_list_of_table_values(cur.execute(sel_income_docs_br.format(self.start_data, self.end_data)),cur.description)

        conn.commit()
        conn.close()
        return income_pays_list, out_docs_list, out_pays_list, income_docs_list
        pass

    def contragent_ops_result(self, income_list):
        result = []
        for contragent in income_list:
            result +=[{'name': contragent[0]['contragent_name'], 'parent':contragent[0]['parent'],  'sum':round(sum([float(x['summ']) for x in contragent]),2)}]
        return result
        pass
        

    def found_result(self, one_list, sec_list):
        out_list = []
        in_list = []
        for doc in one_list:
            for ops in sec_list:
                if doc['parent'] == ops['parent']:
                    if doc['sum'] > ops['sum']:
                        out_list+=[{'name':doc['name'], 'summ':round(doc['sum'] -  ops['sum'],2)}]

                    if doc['sum'] < ops['sum']:
                        in_list+=[{'name':doc['name'], 'summ': round(ops['sum'] - doc['sum'],2)}]

        return(out_list, in_list)           
        pass


    def show_contragent_balance(self):
        out_pays = self.contragent_ops_result(grouping_by_key(self.get_ops_list()[2],'parent'))
        income_docs= self.contragent_ops_result(grouping_by_key(self.get_ops_list()[3], 'parent'))

        income_pays = self.contragent_ops_result(grouping_by_key(self.get_ops_list()[0], 'parent'))
        outcome_docs= self.contragent_ops_result(grouping_by_key(self.get_ops_list()[1], 'parent'))

        provider_debt = self.found_result(out_pays, income_docs)[0]
        provider_prepay = self.found_result(out_pays, income_docs)[1]

        buyer_debt =self.found_result(income_pays, outcome_docs)[0]
        buyer_prepay = self.found_result(income_pays, outcome_docs)[1]

        return provider_debt, provider_prepay, buyer_debt, buyer_prepay
        pass


    def create_hvosty_excel(self):
        output_doc = BASE_DIR+'\\'+"resultat.xlsx"
        openpyxl.Workbook().save(output_doc)
        output_list = load_workbook(output_doc,data_only = True)
        main_out_sheet = output_list.active


        def insert_cell(row_val, col_val, cell_value):
            main_out_sheet.cell(row = row_val, column = col_val).value = cell_value
            pass

        insert_cell(1, 1, "Контрагент")
        insert_cell(1, 2, "Сумма")

        first_space = len(self.show_contragent_balance()[0])+12
        sec_space =  first_space++len(self.show_contragent_balance()[1])+6
        thr_space = sec_space+len(self.show_contragent_balance()[2])+6

        insert_cell(3, 1, 'АВАНСЫ ПОСТАВЩИКАМ(ДЕБ 60)'  )
        for i in range(len(self.show_contragent_balance()[0])):
            insert_cell(i+5, 1, self.show_contragent_balance()[0][i]['name'] )
            insert_cell(i+5, 2, self.show_contragent_balance()[0][i]['summ'])


        insert_cell(first_space, 1, 'ЗАДОЛЖЕННОСТЬ ПЕРЕД ПОСТАВЩИКАМИ(КР 60)'  )
        for i in range(len(self.show_contragent_balance()[1])):
            insert_cell(i+first_space+2, 1, self.show_contragent_balance()[1][i]['name'] )
            insert_cell(i+first_space+2, 2, self.show_contragent_balance()[1][i]['summ'] )    

        insert_cell(sec_space, 1, 'АВАНСЫ ПОКУПАТЕЛЕЙ (КР 62 ОПЛАЧЕН0, НЕ ОТГРУЖЕНО)' )
        for i in range(len(self.show_contragent_balance()[2])):
            insert_cell(i+sec_space+2, 1, self.show_contragent_balance()[2][i]['name'] )
            insert_cell(i+sec_space+2, 2, self.show_contragent_balance()[2][i]['summ'] )  
        
        insert_cell(thr_space, 1, 'ЗАДОЛЖЕННОСТЬ ПОКУПАТЕЛЕЙ (Д 62)' )
        for i in range(len(self.show_contragent_balance()[3])):
            insert_cell(i+thr_space+2, 1, self.show_contragent_balance()[3][i]['name'] )
            insert_cell(i+thr_space+2, 2, self.show_contragent_balance()[3][i]['summ'] ) 
        
        output_list.save(filename = output_doc)

        return(str(output_doc))
        pass


#------------------------SVERKA SS PORTALOM--------------------
class PortalDifference:
    def __init__(self,  data_start=None,  data_end=None,  base_name = None, request_type = None, doc_name = None):

        self.data_start = data_start
        self.data_end = data_end
        self.base_name = base_name
        self.request_type = request_type
        self.doc_name = doc_name


    def nds_docs_list(self):
        conn = sqlite3.connect(self.base_name)
        cur = conn.cursor()
        doc_list = []
        commands =[]
        if self.request_type == "исходящий":
            commands = [income_tn_nds, income_serv_nds, income_tovary]

        if self.request_type == "входящий":
            commands = [outcome_tn_nds, outcome_serv_nds]

        for command in commands:
            doc_list += [create_list_of_table_values(cur.execute(command.format(self.data_start, self.data_end)),cur.description)]
        conn.commit()
        conn.close()

        if len(doc_list) == 2:
            doc_list = doc_list[0]+doc_list[1]
        elif len(doc_list) == 3:
            doc_list = doc_list[0]+doc_list[1]+doc_list[2]  
        return doc_list
        pass


    def get_eschf_data(self):

        first_list_from_excel =[]
        portal_list = load_workbook(self.doc_name,data_only = True)
        main_inner_sheet = portal_list.active

        if self.request_type == "исходящий":
            unp_number = 9
            name_col = 11
            list_number = 0

        if self.request_type == "входящий":
            unp_number = 2
            name_col = 4
            list_number = 5 

        for x in range(1, 8):
            if main_inner_sheet.cell(row=x, column=1).value == "Код страны поставщика": 
                start_point = x+1

        for i in range(start_point, main_inner_sheet.max_row+1):
            if main_inner_sheet.cell(row=i, column=18).value != "Аннулирован": 
                if main_inner_sheet.cell(row=i, column=unp_number).value != None: 
                            first_list_from_excel += [{ 
                            "unp" : main_inner_sheet.cell(row=i, column=unp_number).value, 
                            "name" : main_inner_sheet.cell(row=i, column=name_col).value,
                            "nds" : main_inner_sheet.cell(row=i, column=42).value,
                            }]

        return create_sorted_list(first_list_from_excel)
        pass


    def find_difference(self):
        not_in_excel = self.get_eschf_data()
        not_in_base = self.nds_docs_list()


        for i in self.nds_docs_list():
            for x in self.get_eschf_data():
                if i['unp'] == x['unp'] and i['nds'] == x['nds']:
                    not_in_excel.remove(x)
                    not_in_base.remove(i)
        return(not_in_excel,not_in_base)
        pass


    def insert_into_excel(self):

        output_doc = BASE_DIR+"result.xlsx"
        openpyxl.Workbook().save(output_doc)
        output_list = load_workbook(output_doc,data_only = True)
        main_out_sheet = output_list.active

        def insert_cell(row_val, col_val, cell_value):
            main_out_sheet.cell(row = row_val, column = col_val).value = cell_value
            pass

        for x in (1,4):
            insert_cell(3, x, "Контрагент")

        for i in (2,5):
            insert_cell(3, i, "НДС")    

        insert_cell(1, 4, "Есть в базе, нет на портале")
        insert_cell(1, 1, "Есть на портале, нет в базе")

        insert_cell(2, 4, "Весь НДС из базы")
        insert_cell(2, 1, "Весь НДС с Портала")


        for i in range(len(self.find_difference()[0])):
            insert_cell(i+5, 1, self.find_difference()[0][i]['name'])
            insert_cell(i+5, 2, self.find_difference()[0][i]['nds'])

        for i in range(len(self.find_difference()[1])):
            insert_cell(i+5, 4, self.find_difference()[1][i]['name'])
            insert_cell(i+5, 5, self.find_difference()[1][i]['nds'])

        output_list.save(filename = output_doc)
        return(str(output_doc))
        pass

#-------STATISTICA-----------------------------


def get_today_course():
    courses_list = []
    for i in [requests.get(nbrb_rates_today.format(x['code_nbrb'])).json() for x in  rates]:
        courses_list += [{"cur_name":i["Cur_Name"],"cur_scale":i["Cur_Scale"],"cur_rate":i["Cur_OfficialRate"] }]
    return courses_list
    pass


class CurrencyStat:
    def __init__(self, 
                select_command=None, 
                data_start=None, 
                data_end=None, 
                base_name = None,
                request_type = None,):

       self.select_command = select_command
       self.data_start = data_start
       self.data_end = data_end
       self.base_name = base_name
       self.request_type = request_type

    def create_rates_list(self, select_curr):
        conn = sqlite3.connect(TO_BASE_PATH+'sqlite_bases'+'\\'+'courses.sqlite')
        cur = conn.cursor()
        sql_request = select_curr.format("'"+str(self.data_start)+"'", "'"+str(self.data_end)+"'")
        return create_list_of_table_values(cur.execute(sql_request),cur.description)
        pass


    def transform_sql_to_list(self):
        conn = sqlite3.connect(self.base_name)
        cur = conn.cursor()

        if self.request_type ==  "входящий":
            sel_request = select_curr_income
        if self.request_type == "исходящий":
            sel_request = select_curr_outcome

        sql_request = sel_request.format("'"+self.data_start+"'","'"+self.data_end+"'")


        return create_list_of_table_values(cur.execute(sql_request),cur.description)
        pass

    def result(self):
        eur = []
        usd = []
        rub = []
        for x in self.transform_sql_to_list():
            if x['currency_type'] == '3':
                for y in self.create_rates_list(select_eur_course):
                    for i in self.create_rates_list(select_usd_course):
                        if x['doc_date'] == y['data'] and x['doc_date']== i['data'] and x['name'] != None and type(x['name']) != None:
                            eur += [{
                                    'doc': x['document_name'],  
                                    'date' : x['doc_date'], 
                                    'contragent':x['contragent_name'],
                                    'country': x['name'], 
                                    'summ':x['summ'], 
                                    'rate_on_date':y['rate'],
                                    'usd_rate_on_date':i['rate'],
                                    'bel_sum':round(float(x['summ'])*float(y['rate']),3),
                                    'usd_sum':round(float(x['summ'])*float(y['rate'])/float(i['rate']),3),
                                    'curr_name':'EUR',
                                    }]

            if x['currency_type'] == '7':
                for y in self.create_rates_list(select_usd_course):
                    for i in self.create_rates_list(select_usd_course):
                        if x['doc_date'] == y['data'] and x['doc_date']== i['data'] and x['name'] != None and type(x['name']) != None:
                            usd += [{
                                    'doc': x['document_name'],  
                                    'date' : x['doc_date'], 
                                    'contragent':x['contragent_name'],
                                    'country': x['name'], 
                                    'summ':x['summ'], 
                                    'rate_on_date':y['rate'],
                                    'usd_rate_on_date':i['rate'],
                                    'bel_sum':round(float(x['summ'])*float(y['rate']),3),
                                    'usd_sum':round(float(x['summ'])*float(y['rate'])/float(i['rate']),3),
                                    'curr_name':'USD',
                                    }]


            if x['currency_type'] == '6':
                for y in self.create_rates_list(select_rus_course):
                    for i in self.create_rates_list(select_usd_course):
                        if x['doc_date'] == y['data'] and x['doc_date']== i['data'] and x['name'] != None  and type(x['name']) != None:
                            rub += [{
                                    'doc': x['document_name'],  
                                    'date' : x['doc_date'], 
                                    'contragent':x['contragent_name'], 
                                    'country': x['name'], 
                                    'summ':x['summ'], 
                                    'rate_on_date':y['rate'],
                                    'usd_rate_on_date':i['rate'],
                                    'bel_sum':round(float(x['summ'])*float(y['rate']/float(y['scale'])),3),
                                    'usd_sum':round(float(x['summ'])*float(y['rate']/float(y['scale'])),3)/float(i['rate']),
                                    'curr_name':'RUB',
                                    }]
        return (eur, usd, rub)
        pass


    def stat_for_country(self):
        state_for_cuntry = []
        
        for x in grouping_by_key(self.result()[0]+self.result()[1]+self.result()[2], 'country'):
            state_for_cuntry+=[{
                    'country':x[0]['country'],
                    'summ':round(sum([float(n['summ']) for n in  x]),2),
                    'bel_sum':round(sum([float(n['bel_sum']) for n in  x]),2),
                    'usd_sum':round(sum([float(n['usd_sum']) for n in  x]),2),
                    }]
        return state_for_cuntry
        pass



    def final_grouping(self):
        list_of_lists = []
        final = []
        for count in self.result():
            for i in grouping_by_key(count, 'country'):
                for x in grouping_by_key(i, 'date'):
                    final+=[{
                    'date':x[0]['date'],
                    'contragent':x[0]['contragent'],
                    'country':x[0]['country'],
                    'summ':round(sum([float(n['summ']) for n in  x]),2),
                    'rate_on_date':x[0]['rate_on_date'],
                    'usd_rate_on_date':x[0]['usd_rate_on_date'],
                    'bel_sum':round(sum([float(n['bel_sum']) for n in  x]),2),
                    'usd_sum':round(sum([float(n['usd_sum']) for n in  x]),2),
                    'curr_name':x[0]['curr_name'],
                    }]
            list_of_lists+=[final]

        return list_of_lists
        pass


    def create_statistica_excel(self):
        output_doc = BASE_DIR+"result.xlsx"
        openpyxl.Workbook().save(output_doc)
        output_list = load_workbook(output_doc,data_only = True)
        main_out_sheet = output_list.active
        output_list.create_sheet("full_stat")
        full_stat = output_list["full_stat"]


        income_list = self.final_grouping()
        income_list_main = self.stat_for_country()

        def insert_cell(row_val, col_val, cell_value):
            full_stat.cell(row = row_val, column = col_val).value = cell_value
            pass  


        def insert_cell_main(row_val, col_val, cell_value):
            main_out_sheet.cell(row = row_val, column = col_val).value = cell_value
            pass 

        insert_cell(1, 1, "Дата")
        insert_cell(1, 2, "Контрагент")
        insert_cell(1, 3, "Страна")
        insert_cell(1, 4, "Валюта")
        insert_cell(1, 5, "Сумма в валюте")
        insert_cell(1, 6, "Сумма в бел. рублях")
        insert_cell(1, 7, "Сумма в долларах")
        insert_cell(1, 8, "Курс валюты на дату")
        insert_cell(1, 9, "Курс доллара на дату")


        insert_cell_main(1, 1, "Страна")
        insert_cell_main(1, 2, "Сумма в валюте платежа")
        insert_cell_main(1, 3, "Сумма в бел. рублях")
        insert_cell_main(1, 4, "Сумма в долларах")


        for i in range(len(income_list_main)):
            insert_cell_main(i+3, 1, income_list_main[i]['country'])
            insert_cell_main(i+3, 2, income_list_main[i]['summ'])
            insert_cell_main(i+3, 3, income_list_main[i]['bel_sum'])
            insert_cell_main(i+3, 4, income_list_main[i]['usd_sum'])


    
        for i in range(len(income_list[0])):
            insert_cell(i+3, 1, income_list[0][i]['date'])
            insert_cell(i+3, 2, income_list[0][i]['contragent'])
            insert_cell(i+3, 3, income_list[0][i]['country'])
            insert_cell(i+3, 4, income_list[0][i]['curr_name'])
            insert_cell(i+3, 5, income_list[0][i]['summ'])
            insert_cell(i+3, 6, income_list[0][i]['bel_sum'])
            insert_cell(i+3, 7, income_list[0][i]['usd_sum'])
            insert_cell(i+3, 8, income_list[0][i]['rate_on_date'])
            insert_cell(i+3, 9, income_list[0][i]['usd_rate_on_date'])

        output_list.save(filename = output_doc)
        return(str(output_doc))
        pass


##################----------COURSES--------------##########

class CoursesUpdater:

    def today_updater(self):
        conn = sqlite3.connect('courses.sqlite')
        cur = conn.cursor()
        today_course = []
        for val in rates:
            today_course +=  create_list_of_table_values(cur.execute(select_course_on_date.format(val['name'], "'"+str(date.today())+"'")),cur.description)

        datas_table = create_list_of_table_values(cur.execute(select_course_data.format('usd')),cur.description)
        for i in range(len(rates)):
            datas_table = create_list_of_table_values(cur.execute(select_course_data.format(rates[i]['name'])),cur.description)
            if str(date.today()) != datas_table[-1]['data']:
                cur.executemany(insert_courses.format(rates[i]['name']),  [(str(date.today()), get_today_course()[i]['cur_name'],get_today_course()[i]['cur_scale'],get_today_course()[i]['cur_rate'])])    
        conn.commit()
        conn.close()
        return today_course
        pass


from django.core.validators import FileExtensionValidator

class Upload_file(models.Model):
    uploaded_file = models.FileField(upload_to='docs/', validators=[FileExtensionValidator(['xlsx'])])


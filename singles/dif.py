#! /usr/bin/env python
# -*- coding: utf-8 -*
import os
import sqlite3
import openpyxl
from openpyxl import load_workbook,Workbook
import itertools
from itertools import groupby
from operator import itemgetter
from singles import sql_commands as sq_c
from singles import variables as var


BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'


data_start = "2018-01-01"
data_end=  "2018-31-03"

#ВХОДЯЩИЙ:
data_type = 'income'
unp_number = 2
name_col = 4
list_number = 1


#ИСХОДЯЩИЙ
#data_type = 'outcome'
#unp_number = 9
#name_col = 11
#list_number = 0


output_doc = BASE_DIR+'\\'+"result_{}.xlsx".format(data_type)
openpyxl.Workbook().save(output_doc)

output_list = load_workbook(output_doc,data_only = True)
main_out_sheet = output_list.active



def sum_of_list(sum_name,income_list):
	return sum([x[sum_name] for x in income_list])
	pass


def create_sorted_list(income_list):
	output_list = []
	full_grouped_list = []

	sorted_list = sorted(income_list, key=itemgetter('name'))

	for key, group in itertools.groupby(sorted_list, key=lambda x:x['name']):
		grouped_sorted_list = list(group)
		full_grouped_list += [grouped_sorted_list]								

	for i in full_grouped_list:
		output_list+=[{'name': i[0]['name'],
					   'unp':str(i[0]['unp']), 
					   'nds':round(sum([x['nds'] for x in i]),2)}]

	return output_list
	pass



def	get_eschf_data(main_inner_sheet):
	first_list_from_excel =[]

	for x in range(1, 8):
		if main_inner_sheet.cell(row=x, column=1).value == "Код страны поставщика": 
			start_point = x+1

	for i in range(start_point, main_inner_sheet.max_row+1):
		if  main_inner_sheet.cell(row=i, column=18).value != "Аннулирован": 
			if main_inner_sheet.cell(row=i, column=unp_number).value != None: 
				first_list_from_excel += [{ 
						"unp" : main_inner_sheet.cell(row=i, column=unp_number).value, 
						"name" : main_inner_sheet.cell(row=i, column=name_col).value,
						"nds" : main_inner_sheet.cell(row=i, column=42).value,
						}]
	return create_sorted_list(first_list_from_excel)
	pass


def transform_sql_to_list(cursor, request_command, *condition):
    if len(condition) ==2:
        sql_request = request_command.format(condition[0],condition[1])
    if len(condition) ==3:
        sql_request = request_command.format(condition[0],condition[1],condition[2])
    if len(condition) ==4:
        sql_request = request_command.format(condition[0],condition[1],condition[2],condition[3])

    return var.create_list_of_table_values(cursor.execute(sql_request),cursor.description)
    pass


def curent_finace_states(start, end, cursor):
	list_vhod_nds_usl = transform_sql_to_list(cursor, sq_c.sel_vhod_usl_with_nds, sq_c.tn_buyers,  "'"+start+"'",  "'"+str(end)+"'" )
	list_vhod_nds_tn = transform_sql_to_list(cursor, sq_c.sel_vhod_tn_with_nds, sq_c.tn_buyers,  "'"+start+"'",  "'"+str(end)+"'" )
	list_tovary_nds  = transform_sql_to_list(cursor, sq_c.sel_tovary_with_vhod_nds,   "'"+start+"'",  "'"+str(end)+"'" )

	list_ishod_nds_usl = transform_sql_to_list(cursor, sq_c.sel_usl_with_ishod_nds, sq_c.tn_providers,  "'"+start+"'",  "'"+str(end)+"'" )
	list_ishod_nds_tn = transform_sql_to_list(cursor, sq_c.sel_tn_with_ishod_nds, sq_c.tn_providers,  "'"+start+"'",  "'"+str(end)+"'" )

	unsorted_full_list_vhod = [i for i in list_tovary_nds+list_vhod_nds_tn+list_vhod_nds_usl if i['nds'] != 0.0 and i['nds']!=None]
	unsorted_full_list_ishod = [i for i in list_ishod_nds_tn+list_ishod_nds_usl if i['nds'] != 0.0 and i['nds']!=None]
	
	list_vhod = create_sorted_list(unsorted_full_list_vhod)
	list_ishod = create_sorted_list(unsorted_full_list_ishod)

	return (list_ishod, list_vhod)
	pass

def find_difference(main_inner_sheet, sql_base_name):
	connect = sqlite3.connect(sql_base_name )
	cursor = connect.cursor()
	not_in_excel = get_eschf_data(main_inner_sheet)
	not_in_base = curent_finace_states(data_start, data_end, cursor)[list_number]


	for i in curent_finace_states(data_start, data_end, cursor)[list_number]:
		for x in get_eschf_data(main_inner_sheet):
			if i['unp'] == x['unp'] and i['nds'] == x['nds']:
				not_in_excel.remove(x)
				not_in_base.remove(i)
						
	connect.commit()
	connect.close()
	return(not_in_excel,not_in_base)
	pass



def insert_into_excel(request, excel_income,base_name):

	to_portal_docs_path = BASE_DIR+'docs_from_portal'+'\\'
	portal_list = load_workbook(to_portal_docs_path+excel_income,data_only = True)
	main_inner_sheet = portal_list.active
	sql_base_name = BASE_DIR+'bases'+'\\'+base_name+'.sqlite'
	print(sql_base_name)


	connect = sqlite3.connect(sql_base_name)
	cursor = connect.cursor()

	def insert_cell(row_val, col_val, cell_value):
		main_out_sheet.cell(row = row_val, column = col_val).value = cell_value
		pass

	for x in (1,4):
		insert_cell(3, x, "Контрагент")

	for i in (2,5):
		insert_cell(3, i, "НДС")	

	insert_cell(1, 4, "Есть в базе, нет на портале")
	insert_cell(1, 1, "Есть на портале, нет в базе")

	insert_cell(2, 4, "Весь НДС из базы")
	insert_cell(2, 1, "Весь НДС с Портала")


	insert_cell(2, 5, sum_of_list('nds',curent_finace_states(data_start, data_end, cursor)[list_number]))
	insert_cell(2, 2, sum_of_list('nds', get_eschf_data(main_inner_sheet)))


#	insert_cell(len(find_difference(main_inner_sheet, sql_base_name)[0])+6, 1,"Всего")
#	insert_cell(len(find_difference(main_inner_sheet, sql_base_name)[0])+6, 2,sum_of_list('nds', find_difference(main_inner_sheet, sql_base_name)[0]))


#	insert_cell(len(find_difference(main_inner_sheet, sql_base_name)[1])+6, 4,"Всего")
#	insert_cell(len(find_difference(main_inner_sheet, sql_base_name)[1])+6, 5,	sum_of_list('nds', find_difference(main_inner_sheet, sql_base_name)[1]))


	for i in range(len(find_difference(main_inner_sheet, sql_base_name)[0])):
		insert_cell(i+5, 1, find_difference(main_inner_sheet, sql_base_name)[0][i]['name'])
		insert_cell(i+5, 2, find_difference(main_inner_sheet, sql_base_name)[0][i]['nds'])

	for i in range(len(find_difference(main_inner_sheet, sql_base_name)[1])):

		insert_cell(i+5, 4, find_difference(main_inner_sheet, sql_base_name)[1][i]['name'])
		insert_cell(i+5, 5, find_difference(main_inner_sheet, sql_base_name)[1][i]['nds'])

	output_list.save(filename = output_doc)
	connect.commit()
	connect.close()
	return(str(output_doc))
	pass




#! /usr/bin/env python
# -*- coding: utf-8 -*
import os
import sys
from openpyxl import load_workbook,Workbook



BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'

base_list = BASE_DIR+'transkom_in.xlsx'
wb = load_workbook(base_list)
sheet = wb.active

#https://habrahabr.ru/company/otus/blog/331998/
#http://openpyxl.readthedocs.io/en/default/tutorial.html

def	get_cols_data(col_name,multiplier,sheetname):
		datas = []
		names =[]
		main_sheet = wb["one"]
		for i in range(3,main_sheet.max_row):
			if i >0 and main_sheet.cell(row=i, column=3).value != None:
				names +=[main_sheet.cell(row=i, column=2).value]
				datas +=[round(main_sheet.cell(row=i, column=3).value*multiplier,2)]


		for i in range(len(datas)):
			wb[sheetname].cell(row =i+30, column = 3).value = datas[i]
		wb.save(filename = base_list)		


united_data = get_cols_data(3,0.1984,"United")
mitada_data = get_cols_data(3,0.1522,"Mitada")
bona_data = get_cols_data(3,0.1111,"Bona")



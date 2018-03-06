#! /usr/bin/env python
# -*- coding: utf-8 -*
import dbf
from openpyxl import load_workbook,Workbook


excel_document = 'amedenta.xlsx'

amedenta_list = load_workbook(excel_document,data_only = True)

main_inner_sheet = amedenta_list.active


#names = (name, code, auantity, unit, price, summ, tnvd, weight, weight_summ, country)

def	get_excel_data():
	data_from_excel =[]

	for i in range(6,main_inner_sheet.max_row):
		if main_inner_sheet.cell(row=i, column=3).value != None:
			data_from_excel += [(
								str(main_inner_sheet.cell(row=i, column=3).value),
								str(main_inner_sheet.cell(row=i, column=4).value),
								str(main_inner_sheet.cell(row=i, column=5).value),
								main_inner_sheet.cell(row=i, column=6).value,
								str(main_inner_sheet.cell(row=i, column=7).value),
								str(main_inner_sheet.cell(row=i, column=8).value),
								str(main_inner_sheet.cell(row=i, column=9).value),
								str(main_inner_sheet.cell(row=i, column=10).value),
								str(round(main_inner_sheet.cell(row=i, column=11).value,3)),
								main_inner_sheet.cell(row=i, column=12).value,
								)]

	return data_from_excel
	pass

data = get_excel_data()
table = dbf.Table('SC245')
#print(table.codepage)
#print(data[5])
table.open()
#m = str('сука!!!!'.encode('ascii', 'ignore').decode('cp1251'))
m= str('Ретрактор  для губ и щек'.encode('cp1251'))
print(m)
#cp437. 866, 1251, utf8,koi8-r,iso-8859-1. cp1252

for datum in( ('1VCIB','0','1532','KerrHawe Polishing Kit '),  ):
    table.append(datum)



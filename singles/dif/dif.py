#! /usr/bin/env python
# -*- coding: utf-8 -*
import os
import sys
import sqlite3
from openpyxl import load_workbook,Workbook
import itertools
from itertools import groupby
from operator import itemgetter
path_to_file = os.path.dirname(os.path.abspath(__file__))+'\\'
convert_to_list = path_to_file.split('\\')[:-2]
root_path = '\\'.join(convert_to_list)
sys.path.append(root_path)
import sql_commands as sq_c
import variables as var


BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'
sql_base_name = 'avangard.sqlite'
conn = sqlite3.connect(sql_base_name )
cur = conn.cursor()


excel_income = 'avangard_income_jan_feb.xlsx'
excel_outcome = 'result.xlsx'

excel_sheet_name = "avangard_income_jan_feb"
output_sheet_name = "avangard_result"

data_start = "2018-01-01"
data_end=  "2018-02-28"


#ВХОДЯЩИЙ:
unp_number = 2
name_col = 4
list_number = 1


#ИСХОДЯЩИЙ
#unp_number = 9
#name_col = 11
#list_number = 0



portal_doc = BASE_DIR+excel_income
portal_list = load_workbook(portal_doc,data_only = True)
main_inner_sheet = portal_list[excel_sheet_name]


output_doc = BASE_DIR+excel_outcome
output_list = load_workbook(output_doc,data_only = True)
main_out_sheet = output_list[output_sheet_name]


def sum_of_list(sum_name,income_list):
	return sum([x[sum_name] for x in income_list])
	pass


def create_sorted_list(income_list):
	output_list = []
	full_grouped_list = []

	sorted_list = sorted(income_list, key=itemgetter('name'))

	for key, group in itertools.groupby(sorted_list, key=lambda x:x['name']):
		grouped_sorted_list = list(group)
		full_grouped_list += [grouped_sorted_list]								

	for i in full_grouped_list:
		output_list+=[{'name': i[0]['name'],'unp':str(i[0]['unp']), 'nds':round(sum([x['nds'] for x in i]),2)}]

	return output_list
	pass


def	get_eschf_data():
	first_list_from_excel =[]
	full_grouped_list = []
	outcoming_list =[]

	for i in range(4,main_inner_sheet.max_row):
		if  main_inner_sheet.cell(row=i, column=18) != "Аннулирован" and main_inner_sheet.cell(row=i, column=unp_number).value != None: 
			first_list_from_excel += [{ 
					"unp" : main_inner_sheet.cell(row=i, column=unp_number).value, 
					"name" : main_inner_sheet.cell(row=i, column=name_col).value,
					"nds" : main_inner_sheet.cell(row=i, column=42).value,
#					"full_sum" : main_inner_sheet.cell(row=i, column=42).value,
					}]

	return create_sorted_list(first_list_from_excel)
	pass


def transform_sql_to_list(cursor, request_command, *condition):
    if len(condition) ==2:
        sql_request = request_command.format(condition[0],condition[1])
    if len(condition) ==3:
        sql_request = request_command.format(condition[0],condition[1],condition[2])
    if len(condition) ==4:
        sql_request = request_command.format(condition[0],condition[1],condition[2],condition[3])

    return var.create_list_of_table_values(cursor.execute(sql_request),cursor.description)
    pass


def curent_finace_states(start, end, cursor):
	list_vhod_nds_usl = transform_sql_to_list(cursor, sq_c.sel_vhod_usl_with_nds, sq_c.tn_buyers,  "'"+start+"'",  "'"+str(end)+"'" )
	list_vhod_nds_tn = transform_sql_to_list(cursor, sq_c.sel_vhod_tn_with_nds, sq_c.tn_buyers,  "'"+start+"'",  "'"+str(end)+"'" )
	list_tovary_nds  = transform_sql_to_list(cursor, sq_c.sel_tovary_with_vhod_nds,   "'"+start+"'",  "'"+str(end)+"'" )

	list_ishod_nds_usl = transform_sql_to_list(cursor, sq_c.sel_usl_with_ishod_nds, sq_c.tn_providers,  "'"+start+"'",  "'"+str(end)+"'" )
	list_ishod_nds_tn = transform_sql_to_list(cursor, sq_c.sel_tn_with_ishod_nds, sq_c.tn_providers,  "'"+start+"'",  "'"+str(end)+"'" )

	unsorted_full_list_vhod = [i for i in list_tovary_nds+list_vhod_nds_tn+list_vhod_nds_usl if i['nds'] != 0.0 and i['nds']!=None]
	unsorted_full_list_ishod = [i for i in list_ishod_nds_usl+list_ishod_nds_tn if i['nds'] != 0.0 and i['nds']!=None]
	
	list_vhod = create_sorted_list(unsorted_full_list_vhod)
	list_ishod = create_sorted_list(unsorted_full_list_ishod)

	return (list_ishod, list_vhod)
	pass


def find_difference():
	not_in_excel = get_eschf_data()
	not_in_base = curent_finace_states(data_start, data_end, cur)[list_number]

	final_not_in_portal = []
	final_not_in_base = []
	govno = []
	mocha = []

	for i in curent_finace_states(data_start, data_end, cur)[list_number]:
		for x in get_eschf_data():
			if i['unp'] == x['unp'] and i['nds'] == x['nds']:
				not_in_excel.remove(x)
				not_in_base.remove(i)
						

	return(not_in_excel,not_in_base)
	pass

find_difference()


def insert_into_excel():

	def insert_cell(row_val, col_val, cell_value):
		main_out_sheet.cell(row = row_val, column = col_val).value = cell_value
		pass

	insert_cell(2, 1, "Контрагент")
	insert_cell(2, 4, "Контрагент")

	insert_cell(2, 2, "НДС")
	insert_cell(2, 5, "НДС")

	insert_cell(1, 1, "Весь НДС из базы")
	insert_cell(1, 3, "Весь НДС с Портала")


	insert_cell(1, 2, sum_of_list('nds',curent_finace_states(data_start, data_end, cur)[list_number]))
	insert_cell(1, 4, sum_of_list('nds', get_eschf_data()))


	insert_cell(len(find_difference()[0])+4, 1,"Всего")
	insert_cell(len(find_difference()[0])+4, 2,sum_of_list('nds', find_difference()[0]))


	insert_cell(len(find_difference()[1])+4, 4,"Всего")
	insert_cell(len(find_difference()[1])+4, 5,	sum_of_list('nds', find_difference()[1]))


	for i in range(len(find_difference()[0])):
		insert_cell(i+3, 1, find_difference()[0][i]['name'])
		insert_cell(i+3, 2, find_difference()[0][i]['nds'])

	for i in range(len(find_difference()[1])):

		insert_cell(i+3, 4, find_difference()[1][i]['name'])
		insert_cell(i+3, 5, find_difference()[1][i]['nds'])

	output_list.save(filename = output_doc)
	pass


#insert_into_excel()
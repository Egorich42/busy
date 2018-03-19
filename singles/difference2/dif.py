#! /usr/bin/env python
# -*- coding: utf-8 -*
import os
import sys

path_to_file = os.path.dirname(os.path.abspath(__file__))+'\\'
convert_to_list = path_to_file.split('\\')[:-2]
root_path = '\\'.join(convert_to_list)
sys.path.append(root_path)
import sql_commands as sq_c
import variables as var


import sqlite3

from openpyxl import load_workbook,Workbook
import itertools
from itertools import groupby
from operator import itemgetter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'

portal_doc = BASE_DIR+'hpo_in.xlsx'


portal_list = load_workbook(portal_doc,data_only = True)


conn = sqlite3.connect('7.sqlite')
cur = conn.cursor()

main_inner_sheet = portal_list["jan"]


def	get_eschf_data():
	first_list_from_excel =[]
	full_grouped_list = []
	outcoming_list =[]

	for i in range(4,main_inner_sheet.max_row):
		if  main_inner_sheet.cell(row=i, column=18) != "Аннулирован" and main_inner_sheet.cell(row=i, column=2).value != None: 
			first_list_from_excel += [{ 
					"unp" : main_inner_sheet.cell(row=i, column=2).value, 
					"contragent_name" : main_inner_sheet.cell(row=i, column=4).value,
					"nds" : main_inner_sheet.cell(row=i, column=41).value,
					"full_sum" : main_inner_sheet.cell(row=i, column=42).value,
					}]


	sorted_list_from_excel = sorted(first_list_from_excel, key=itemgetter('contragent_name'))

	for key, group in itertools.groupby(sorted_list_from_excel, key=lambda x:x['contragent_name']):
		grouped_list_of_lists = list(group)
		full_grouped_list+=[grouped_list_of_lists]								


	for i in full_grouped_list:
		outcoming_list+=[{'name': i[0]['contragent_name'],'unp':str(i[0]['unp']), 'nds':round(sum([x['nds'] for x in i]),2)}]



	return outcoming_list
	pass


def transform_sql_to_list(cursor, request_command, *condition):
    if len(condition) ==2:
        sql_request = request_command.format(condition[0],condition[1])
    if len(condition) ==3:
        sql_request = request_command.format(condition[0],condition[1],condition[2])
    if len(condition) ==4:
        sql_request = request_command.format(condition[0],condition[1],condition[2],condition[3])
    
    table = var.create_list_of_table_values(cursor.execute(sql_request),cursor.description)
    return table
    pass


def curent_finace_states(start, end, cursor):
	full_grouped_list = []
	outcoming_list = []

	list_vhod_nds_usl = transform_sql_to_list(cursor, sq_c.sel_vhod_usl_with_nds, sq_c.tn_buyers,  "'"+start+"'",  "'"+str(end)+"'" )
	list_vhod_nds_tn = transform_sql_to_list(cursor, sq_c.sel_vhod_tn_with_nds, sq_c.tn_buyers,  "'"+start+"'",  "'"+str(end)+"'" )
	list_tovary_nds  = transform_sql_to_list(cursor, sq_c.sel_tovary_with_vhod_nds,   "'"+start+"'",  "'"+str(end)+"'" )

	unsorted_full_list = [i for i in list_tovary_nds+list_vhod_nds_tn+list_vhod_nds_usl if i['nds'] != 0.0 and i['nds']!=None]

	sorted_list_from_base = sorted(unsorted_full_list, key=itemgetter('name'))


	for key, group in itertools.groupby(sorted_list_from_base, key=lambda x:x['name']):
		grouped_list_of_lists = list(group)
		full_grouped_list+=[grouped_list_of_lists]								
							

	for i in full_grouped_list:
		outcoming_list+=[{'name': i[0]['name'],'unp':i[0]['unp'], 'nds':round(sum([x['nds'] for x in i]),2)}]

	return outcoming_list
	pass

data1 = curent_finace_states("2017-11-01", "2017-11-31", cur)
data2 = get_eschf_data()

def find_difference():
	not_in_excel = get_eschf_data()
	not_in_base = curent_finace_states("2017-11-01", "2017-11-31", cur)


	for i in curent_finace_states("2017-11-01", "2017-11-31", cur):
		for x in get_eschf_data():
			if i['unp'] == x['unp']:
				not_in_excel.remove(x)

	return not_in_excel

print(find_difference())	

#https://stackoverflow.com/questions/19755376/getting-the-difference-delta-between-two-lists-of-dictionaries
from functools import reduce
s1 = set(reduce(lambda x, y: x + y, [x.items() for x in data1]))
s2 = set(reduce(lambda x, y: x + y, [x.items() for x in data2]))

s2.difference(s1)
s2.symmetric_difference(s1)
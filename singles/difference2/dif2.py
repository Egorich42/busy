#! /usr/bin/env python
# -*- coding: utf-8 -*
import os
import sys
path_to_file = os.path.dirname(os.path.abspath(__file__))+'\\'
convert_to_list = path_to_file.split('\\')[:-2]
root_path = '\\'.join(convert_to_list)
sys.path.append(root_path)
import sql_commands as sq_c
import variables as var
from dbfread import DBF

import sqlite3


from openpyxl import load_workbook,Workbook
import itertools
from itertools import groupby
from operator import itemgetter

conn = sqlite3.connect('hpo.sqlite')
cur = conn.cursor()


dbf_esch='D:\DATA_SETS\himprom_den\SC37543.DBF'
dbf_contragents='D:\DATA_SETS\himprom_den\SC167.DBF'




def get_data_from_dbf(table_name):
	dataset = list(DBF(table_name, encoding='iso-8859-1'))
	return dataset
	pass

esch_list_dbf = get_data_from_dbf(dbf_esch)
contragents_list_dbf = get_data_from_dbf(dbf_contragents)

def get_data_eschf():
	values_list = []
	for i in esch_list_dbf:	
		values_list += [(
						dict(i)['DESCR'].encode('latin1').decode('cp1251'), 
						dict(i)['PARENTEXT'].replace(" ", ""), 
						dict(i)['SP37497'], 
						dict(i)['SP37498'],
						dict(i)['SP37527'],
						dict(i)['ISMARK'],
						)]
	return values_list	
	pass


def get_data_contragents():
	values_list = []
	for i in contragents_list_dbf:	
		values_list += [(
					dict(i)['ID'].replace(" ", ""), 
					dict(i)['SP137'],
					dict(i)['DESCR'].encode('latin1').decode('cp1251'), 
					dict(i)['SP134'].encode('latin1').decode('cp1251'), 
					dict(i)['ISMARK'],
					)]
	return values_list				
	pass

contragents = get_data_contragents()
eschf_list = get_data_eschf()

"""
cur.execute("CREATE table IF NOT EXISTS contragents (id,unp,contragent_name,full_name, deleted);")
cur.execute("CREATE table IF NOT EXISTS outer_eschf (document_name,parent,full_sum, nds, data, deleted);")


cur.executemany('INSERT INTO contragents VALUES (?,?,?,?,?)', contragents)
cur.executemany('INSERT INTO outer_eschf VALUES (?,?,?,?,?,?)', eschf_list)


conn.commit()
conn.close()
"""


###-EXCEL-PART-------------------------##############


BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'

portal_doc = BASE_DIR+'jan_out.xlsx'

portal_list = load_workbook(portal_doc,data_only = True)

main_inner_sheet = portal_list["jan"]
result_sheet = portal_list["result"]


def get_eschf_from_sql():
	full_grouped_list = []
	outcoming_list = []

	select = sq_c.select_eschf_documents.format("'"+"2018-01-01"+"'","'"+"2018-01-31"+"'")
	transform_to_list = var.create_list_of_table_values(cur.execute(select),cur.description)

	sorted_list_from_sql = sorted(transform_to_list, key=itemgetter('contragent_name'))

	for key, group in itertools.groupby(sorted_list_from_sql, key=lambda x:x['unp']):
		grouped_list_of_listd = list(group)
		full_grouped_list+=[grouped_list_of_listd]								


	for i in full_grouped_list:
		outcoming_list+=[{'contragent_name': i[0]['contragent_name'],'unp':i[0]['unp'] , 'nds':round(sum([x['nds'] for x in i]),2)}]

	
#	print(len(transform_to_list))
#	print(len(outcoming_list))	
	return outcoming_list
	pass




def	get_eschf_from_excel():
	first_list_from_excel =[]
	full_grouped_list = []
	outcoming_list = []

	for i in range(4,main_inner_sheet.max_row):
		if  main_inner_sheet.cell(row=i, column=18).value != "Аннулирован" and main_inner_sheet.cell(row=i, column=9).value != None:
			first_list_from_excel += [{
				    "unp" : main_inner_sheet.cell(row=i, column=9).value, 
					"contragent_name" : main_inner_sheet.cell(row=i, column=11).value,
					"nds" : main_inner_sheet.cell(row=i, column=42).value,
					"full_sum" : main_inner_sheet.cell(row=i, column=43).value,
					}]

	sorted_list_from_excel = sorted(first_list_from_excel, key=itemgetter('contragent_name'))

	for key, group in itertools.groupby(sorted_list_from_excel, key=lambda x:x['unp']):
		grouped_list_of_listd = list(group)
		full_grouped_list+=[grouped_list_of_listd]								


	for i in full_grouped_list:
		outcoming_list+=[{'contragent_name': i[0]['contragent_name'],'unp':i[0]['unp'] , 'nds':round(sum([x['nds'] for x in i]),2)}]

	return outcoming_list
	pass


sql_unjoined = get_eschf_from_sql()
exc_unjoined = get_eschf_from_excel()

M = sql_unjoined+exc_unjoined
def suka():
	atp = get_eschf_from_sql()

	for x in get_eschf_from_sql():
		for y in get_eschf_from_excel():
			if x['unp'] == y['unp']:
				atp.remove(y)
	return atp
print(len(suka()))

print(len(sql_unjoined))
print(len(exc_unjoined))



#nds_sum_sql = sum(item['nds'] for item in sql_unjoined)
#nds_sum_portal = sum(item['nds'] for item in exc_unjoined)


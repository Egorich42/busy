#! /usr/bin/env python
# -*- coding: utf-8 -*
import os
from openpyxl import load_workbook,Workbook

"""
import sys
path_to_file = os.path.dirname(os.path.abspath(__file__))+'\\'
convert_to_list = path_to_file.split('\\')[:-2]
root_path = '\\'.join(convert_to_list)
sys.path.append(root_path)
"""



BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'

base_doc = BASE_DIR+'ertex_in_feb.xlsx'
output_doc = BASE_DIR+'ertex_out.xlsx'

income_list = load_workbook(base_doc,data_only = True)
output_list = load_workbook(output_doc,data_only = True)


def	get_cols_data():
	all_data = []

	prices =[]
	income =[]
	codes_1C =[]

	new_prices=[]
	new_income =[]
	new_codes_1c = []

	changed_prices =[]
	changed_codes_1c =[]
	changed_income =[]

	main_sheet = income_list["feb"]
	output_sheet = output_list["jan"]


	for i in range(6,main_sheet.max_row):
		if main_sheet.cell(row=i, column=10).value != None and main_sheet.cell(row=i, column=10).value != 0 and main_sheet.cell(row=i, column=8).value != None:
			all_data += [ { "code_1C" : main_sheet.cell(row=i, column=6).value, 
							"remains_on_warehouse" : main_sheet.cell(row=i, column=8).value,
							"remains_summ" : main_sheet.cell(row=i, column=9).value,
							"income" : main_sheet.cell(row=i, column=10).value,
							"past_price" : main_sheet.cell(row=i, column=11).value
			           	}]


		if main_sheet.cell(row=i, column=10).value != None and main_sheet.cell(row=i, column=10).value != 0 and main_sheet.cell(row=i, column=8).value == None:
			new_prices +=[main_sheet.cell(row=i, column=11).value]
			new_income += [main_sheet.cell(row=i, column=10).value]
			new_codes_1c += [main_sheet.cell(row=i, column=6).value]



	for i in all_data:
		if i["remains_summ"]/i["remains_on_warehouse"] != i["past_price"]:
			changed_codes_1c +=[i["code_1C"]]
			changed_prices += [round(i["remains_summ"]/i["remains_on_warehouse"],2)]
			changed_income += [i["income"]]
		else:
			codes_1C +=[i["code_1C"]]
			prices += [round(i["past_price"],2)]
			income += [i["income"]]

		for i in range(len(changed_codes_1c)):
			output_sheet.cell(row = i+3, column = 1).value = changed_codes_1c[i]
			output_sheet.cell(row = i+3, column = 2).value = changed_income[i]
			output_sheet.cell(row = i+3, column = 3).value = changed_prices[i]
		

	none_error_codes = codes_1C+new_codes_1c
	none_error_income = income+new_income
	none_error_prices = prices+new_prices


	for i in range(len(none_error_codes)):
		output_sheet.cell(row = i+3, column = 5).value = none_error_codes[i]
		output_sheet.cell(row = i+3, column = 6).value = none_error_income[i]
		output_sheet.cell(row = i+3, column = 7).value = none_error_prices[i]


	output_list.save(filename = output_doc)


get_cols_data()

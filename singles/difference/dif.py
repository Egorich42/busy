#! /usr/bin/env python
# -*- coding: utf-8 -*
import os
import sys
from openpyxl import load_workbook,Workbook
import itertools
from itertools import groupby
from operator import itemgetter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'

portal_doc = BASE_DIR+'portal.xlsx'
nds_uslugi_doc = BASE_DIR+'nds_uslugi.xlsx'
nds_tovary_doc = BASE_DIR+'nds_tovary.xlsx'


portal_list = load_workbook(portal_doc,data_only = True)
nds_uslugi_list = load_workbook(nds_uslugi_doc,data_only = True)
nds_tovary_list = load_workbook(nds_tovary_doc,data_only = True)

main_inner_sheet = portal_list["jan"]
uslugi_sheet = nds_uslugi_list.active
tovary_sheet = nds_tovary_list.active



def	get_eschf_data():
	first_list_from_excel =[]
	full_grouped_list =[]
	outcoming_list =[]

	for i in range(4,main_inner_sheet.max_row):
		if  main_inner_sheet.cell(row=i, column=18) != "Аннулирован":
			first_list_from_excel += [{ "unp" : main_inner_sheet.cell(row=i, column=2).value, 
					"contragent_name" : main_inner_sheet.cell(row=i, column=4).value,
					"nds" : main_inner_sheet.cell(row=i, column=42).value,
					}]


	sorted_list_from_excel = sorted(first_list_from_excel, key=itemgetter('unp'))

	for key, group in itertools.groupby(sorted_list_from_excel, key=lambda x:x['unp']):
		grouped_list_of_listd = list(group)
		full_grouped_list+=[grouped_list_of_listd]


	for i in full_grouped_list:
		outcoming_list+=[{'contragent_name': i[0]['contragent_name'], 'nds':round(sum([x['nds'] for x in i]),2)}]


	return outcoming_list
	pass




def	get_excel_data():
	uslugi_from_excel =[]
	tovary_from_excel = []
	full_grouped_list = []
	outcoming_list =[]

	for i in range(7,uslugi_sheet.max_row):
		if uslugi_sheet.cell(row=i, column=1).value != None:
			uslugi_from_excel += [{"contragent_name" : uslugi_sheet.cell(row=i, column=1).value,
									"nds" : uslugi_sheet.cell(row=i, column=2).value,}]


	dicts_from_tovary =[]
	for i in range(7,tovary_sheet.max_row):
		if tovary_sheet.cell(row=i, column=1).value != None:
			tovary_from_excel += [{"contragent_name" : tovary_sheet.cell(row=i, column=1).value,
									"nds" : tovary_sheet.cell(row=i, column=2).value,}]

			
	sorted_list_from_excel = sorted(uslugi_from_excel+tovary_from_excel, key=itemgetter('contragent_name'))

	for key, group in itertools.groupby(sorted_list_from_excel, key=lambda x:x['contragent_name']):
		grouped_list_of_listd = list(group)
		full_grouped_list+=[grouped_list_of_listd]								


	for i in full_grouped_list:
		outcoming_list+=[{'contragent_name': i[0]['contragent_name'], 'nds':round(sum([x['nds'] for x in i]),2)}]



	return outcoming_list
	pass
	


def not_in_base():
	result = get_eschf_data()

	for i in get_excel_data():
		for k in range(len(get_eschf_data())):
			if get_eschf_data()[k]['nds'] == i['nds']:
				result.remove(get_eschf_data()[k])

	
	for x in range(len(result)):
		portal_list["not_in_base"].cell(row=2, column=1).value = "НЕТ В БАЗЕ"
		portal_list["not_in_base"].cell(row=x+3, column=1).value = result[x]['contragent_name']
		portal_list["not_in_base"].cell(row=x+3, column=2).value = result[x]['nds']
		pass	

	portal_list.save(filename = portal_doc)	






def not_in_portal():
	result = get_excel_data()

	for i in get_eschf_data():
		for k in range(len(get_excel_data())):
			if get_excel_data()[k]['nds'] == i['nds']:
				result.remove(get_excel_data()[k])	

	portal_list["not_in_base"].cell(row=2, column=5).value = "НЕТ НА ПОРТАЛЕ"
	for x in range(len(result)):
		portal_list["not_in_base"].cell(row=x+3, column=5).value = result[x]['contragent_name']
		portal_list["not_in_base"].cell(row=x+3, column=6).value = result[x]['nds']
		pass	

	portal_list.save(filename = portal_doc)					









not_in_portal()
not_in_base()

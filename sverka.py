import sqlite3
import itertools
import os

import openpyxl
from openpyxl import load_workbook,Workbook

from itertools import groupby
from collections import defaultdict
from operator import itemgetter



from TESTO import create_list_of_table_values, grouping_by_key, sum_of_list
from TESTO import income_tn_nds, income_serv_nds, income_tovary, outcome_tn_nds, outcome_serv_nds


BASE_DIR = os.path.dirname(os.path.abspath(__file__))+'\\'


def create_sorted_list(income_list):
    output_list = []
    full_grouped_list = []

    sorted_list = sorted(income_list, key=itemgetter('name'))

    for key, group in itertools.groupby(sorted_list, key=lambda x:x['name']):
        grouped_sorted_list = list(group)
        full_grouped_list += [grouped_sorted_list]                              

    for i in full_grouped_list:
        output_list+=[{'name': i[0]['name'],
                       'unp':str(i[0]['unp']), 
                       'nds':round(sum([float(x['nds']) for x in i]),2)}]

    return output_list
    pass

class PortalDifference:
	def __init__(self, 
				data_start=None, 
				data_end=None, 
				base_name = None,
				request_type = None,
				nalog_system = None):

				self.data_start = data_start
				self.data_end = data_end
				self.base_name = base_name
				self.request_type = request_type
				self.nalog_system = nalog_system


	def nds_docs_list(self):
		conn = sqlite3.connect(self.base_name)
		cur = conn.cursor()
		doc_list = []
		commands =[]
		if self.request_type == "исходящий":
			commands = [income_tn_nds, income_serv_nds, income_tovary]

		if self.request_type == "входящий":
			commands = [outcome_tn_nds, outcome_serv_nds]

		for command in commands:
			doc_list += [create_list_of_table_values(cur.execute(command.format(self.data_start, self.data_end)),cur.description)]
		conn.commit()
		conn.close()

		if len(doc_list) == 2:
			doc_list = doc_list[0]+doc_list[1]
		elif len(doc_list) == 3:
			doc_list = doc_list[0]+doc_list[1]+doc_list[2]	
		return doc_list
		pass


	def get_eschf_data(self):
		document = BASE_DIR+'hpo_march_18.xlsx'

		first_list_from_excel =[]
		portal_list = load_workbook(document,data_only = True)
		main_inner_sheet = portal_list.active


		if self.request_type == "исходящий":
			unp_number = 9
			name_col = 11
			list_number = 0

		if self.request_type == "входящий":
			unp_number = 2
			name_col = 4
			list_number = 5	



		for x in range(1, 8):
			if main_inner_sheet.cell(row=x, column=1).value == "Код страны поставщика": 
				start_point = x+1

		for i in range(start_point, main_inner_sheet.max_row+1):
			if main_inner_sheet.cell(row=i, column=18).value != "Аннулирован": 
				if main_inner_sheet.cell(row=i, column=unp_number).value != None: 
							first_list_from_excel += [{ 
							"unp" : main_inner_sheet.cell(row=i, column=unp_number).value, 
							"name" : main_inner_sheet.cell(row=i, column=name_col).value,
							"nds" : main_inner_sheet.cell(row=i, column=42).value,
							}]

		return create_sorted_list(first_list_from_excel)
		pass




	def find_difference(self):
		not_in_excel = self.get_eschf_data()
		not_in_base = self.nds_docs_list()


		for i in self.nds_docs_list():
			for x in self.get_eschf_data():
				if i['unp'] == x['unp'] and i['nds'] == x['nds']:
					not_in_excel.remove(x)
					not_in_base.remove(i)
		return(not_in_excel,not_in_base)
		pass

	def insert_into_excel(self):

		output_doc = BASE_DIR+"result.xlsx"
		openpyxl.Workbook().save(output_doc)
		output_list = load_workbook(output_doc,data_only = True)
		main_out_sheet = output_list.active



		def insert_cell(row_val, col_val, cell_value):
			main_out_sheet.cell(row = row_val, column = col_val).value = cell_value
			pass


		for i in range(len(self.find_difference()[0])):
			insert_cell(i+5, 1, self.find_difference()[0][i]['name'])
			insert_cell(i+5, 2, self.find_difference()[0][i]['nds'])

		for i in range(len(self.find_difference()[1])):
			insert_cell(i+5, 4, self.find_difference()[1][i]['name'])
			insert_cell(i+5, 5, self.find_difference()[1][i]['nds'])

		output_list.save(filename = output_doc)
		return(str(output_doc))
		pass


print(PortalDifference(base_name ="avangard.sqlite", 
				data_start = "'"+ '2018-01-30'+"'", 
				data_end="'"+'2018-02-05'+"'",
				request_type = "исходящий").insert_into_excel())